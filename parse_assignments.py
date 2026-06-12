#!/usr/bin/env python3
"""
Parse '2026 Official Docent Assignment Doc.xlsx' into schedule_import.csv
for the CMOA Docent Schedule Google Sheet.

Output columns match the Schedule tab:
SlotId, Date, Time, TourType, DocentsNeeded, Status, Assigned, Details,
TourLeadSchool, ParticipantSchool, MindfulWelcomeDesk, MindfulTourLead,
DocentsNeeded_Desk, DocentsNeeded_MindfulTour
"""

import openpyxl
import csv
import re
from datetime import datetime, date

wb = openpyxl.load_workbook('2026 Official Docent Assignment Doc.xlsx', data_only=True)

CUTOFF = date(2026, 6, 8)  # include from current week onward

# Name normalization: schedule names -> docent list names
NAME_MAP = {
    'Debby Prence': 'Deborah Prence',
    'Lori Mitchell-McMahon': 'Lori McMahon',
}

def normalize_name(name):
    """Normalize docent name to match the Docents tab"""
    if not name:
        return name
    return NAME_MAP.get(name, name)

def parse_time_from_desc(desc):
    """Extract start time from descriptions like 'Daily, 11–11:30 a.m.' or 'Thurs., 6:30–7:30 p.m.'"""
    if not desc:
        return None
    desc = str(desc)
    # Match patterns like "11–11:30 a.m.", "1:30–2:30 p.m.", "6–7:15 p.m."
    m = re.search(r'(\d{1,2})(?::(\d{2}))?[\s\u2013\u2014-]', desc)
    if not m:
        return None
    hour = int(m.group(1))
    minute = int(m.group(2)) if m.group(2) else 0
    # Determine AM/PM
    if 'p.m.' in desc.lower() or 'pm' in desc.lower():
        if hour < 12:
            hour += 12
    elif 'a.m.' in desc.lower() or 'am' in desc.lower():
        if hour == 12:
            hour = 0
    else:
        # No explicit am/pm - guess based on hour
        if hour < 8:
            hour += 12  # assume PM for small hours

    ampm = 'AM' if hour < 12 else 'PM'
    display_hour = hour if hour <= 12 else hour - 12
    if display_hour == 0:
        display_hour = 12
    return f"{display_hour}:{minute:02d} {ampm}"


def parse_time_from_tour_desc(desc):
    """Extract time from tour descriptions like 'Tour 1: Group Name, 11:00-12:00, ...' or 'CI Activation 12-2'"""
    if not desc:
        return None
    desc = str(desc)

    # CI Activation patterns
    m = re.search(r'CI Activation\s+(\d{1,2})(?::(\d{2}))?[\s\u2013-]', desc)
    if m:
        hour = int(m.group(1))
        minute = int(m.group(2)) if m.group(2) else 0
        if hour < 8: hour += 12
        ampm = 'AM' if hour < 12 else 'PM'
        dh = hour if hour <= 12 else hour - 12
        if dh == 0: dh = 12
        return f"{dh}:{minute:02d} {ampm}"

    # MINDFUL MUSEUM patterns - take first time
    m = re.search(r'MINDFUL MUSEUM.*?(\d{1,2}):(\d{2})', desc)
    if m:
        hour = int(m.group(1))
        minute = int(m.group(2))
        if hour < 8: hour += 12
        ampm = 'AM' if hour < 12 else 'PM'
        dh = hour if hour <= 12 else hour - 12
        if dh == 0: dh = 12
        return f"{dh}:{minute:02d} {ampm}"

    # General time pattern in tour descriptions: "11:00-12:00" or "2pm" or "1pm"
    m = re.search(r'(\d{1,2}):(\d{2})\s*(?:[\s\u2013-]|to)', desc)
    if m:
        hour = int(m.group(1))
        minute = int(m.group(2))
        if hour < 8: hour += 12
        ampm = 'AM' if hour < 12 else 'PM'
        dh = hour if hour <= 12 else hour - 12
        if dh == 0: dh = 12
        return f"{dh}:{minute:02d} {ampm}"

    m = re.search(r'(\d{1,2})\s*(am|pm)', desc, re.I)
    if m:
        hour = int(m.group(1))
        if m.group(2).lower() == 'pm' and hour < 12: hour += 12
        if m.group(2).lower() == 'am' and hour == 12: hour = 0
        ampm = 'AM' if hour < 12 else 'PM'
        dh = hour if hour <= 12 else hour - 12
        if dh == 0: dh = 12
        return f"{dh}:00 {ampm}"

    return None


def clean_docent_name(raw):
    """Clean docent name, removing role annotations like (lead), (shadow), time notations, etc."""
    if not raw:
        return None
    name = str(raw).strip()
    if not name:
        return None
    # Skip non-name entries
    skip_patterns = ['MUSEUM CLOSED', 'LABOR DAY', 'Museum open', 'Juneteenth',
                     'Tour cancelled', 'CANCELLED', 'Daily,', 'Thurs.,', 'Sat.',
                     'Mon.', 'Tue.', 'Wed.', 'Fri.', 'Sun.']
    for pat in skip_patterns:
        if pat.lower() in name.lower():
            return None

    # Remove time annotations and role tags
    name = re.sub(r'\s*\(lead\)', '', name, flags=re.I)
    name = re.sub(r'\s*\(shadow\)', '', name, flags=re.I)
    name = re.sub(r'\s*\d{1,2}:\d{2}(?:\s*,?\s*\d{1,2}:\d{2})*', '', name)
    name = re.sub(r'\s*\d{1,2}:\d{2}', '', name)
    name = name.strip(' ,')

    if not name or len(name) < 3:
        return None
    # Must look like a name (contains a space or known single names)
    if ' ' not in name and name not in ['TBD']:
        return None
    return normalize_name(name)


def is_lead(raw):
    """Check if the entry marks this docent as lead"""
    return bool(raw and re.search(r'\(lead\)', str(raw), re.I))


def is_mm_table(raw):
    """Check if this is a Mindful Museum table role"""
    return bool(raw and re.match(r'MM Table:', str(raw).strip(), re.I))


def is_mm_assist(raw):
    """Check if this is a Mindful Museum assist role"""
    return bool(raw and re.match(r'MM Assist', str(raw).strip(), re.I))


def classify_tour(header):
    """Classify a tour header into a tour type"""
    if not header:
        return None, None
    h = str(header).strip()
    hl = h.lower()

    if 'mindful museum' in hl:
        return 'Mindful Museum', h
    if 'ci activation' in hl:
        return 'CI Activation Tour', h
    if 'tour cancelled' in hl:
        return None, None
    if re.match(r'tour \d+:', hl, re.I):
        # Group/donor/school tour
        # Try to detect school tours vs regular group tours
        if any(w in hl for w in ['school', 'grade', 'students', ' es,', ' hs,', ' ms,']):
            return 'School Tour', h
        return 'Group Tour', h
    if 'donor' in hl:
        return 'Group Tour', h

    return None, None


slots = []
slot_counter = 0

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    # Find all day boundaries (rows where col A has a date)
    day_boundaries = []
    for r in range(1, ws.max_row + 1):
        val = ws.cell(r, 1).value
        if isinstance(val, datetime):
            d = val.date()
            if d >= CUTOFF:
                day_boundaries.append((r, d))

    if not day_boundaries:
        continue

    # Process each day
    for idx, (start_row, tour_date) in enumerate(day_boundaries):
        # End row is the row before the next date, or a reasonable limit
        if idx + 1 < len(day_boundaries):
            end_row = day_boundaries[idx + 1][0] - 1
        else:
            end_row = min(start_row + 30, ws.max_row)

        date_str = tour_date.strftime('%Y-%m-%d')

        # === PERMANENT COLLECTION TOURS (cols C=3, D=4) ===
        # Check if col C in the date row has the PC header
        header_c = ws.cell(start_row, 3).value
        if header_c and 'permanent collection' in str(header_c).lower():
            for r in range(start_row + 1, end_row + 1):
                time_desc = ws.cell(r, 3).value
                docent_name = ws.cell(r, 4).value
                if not time_desc:
                    continue
                time_str = parse_time_from_desc(str(time_desc))
                if not time_str:
                    continue

                name = clean_docent_name(docent_name)
                slot_counter += 1
                sid = f"{date_str}_PC_{slot_counter}"

                is_evening = 'p.m.' in str(time_desc).lower() and any(x in str(time_desc) for x in ['6:', '7:', '6–', '7–'])
                tour_type = 'Permanent Collection Evening' if is_evening else 'Permanent Collection'

                slots.append({
                    'slotId': sid,
                    'date': date_str,
                    'time': time_str,
                    'tourType': tour_type,
                    'docentsNeeded': 1,
                    'status': 'Assigned' if name else 'Open',
                    'assigned': name or '',
                    'details': '',
                    'tourLeadSchool': '',
                    'participantSchool': '',
                    'mindfulWelcomeDesk': '',
                    'mindfulTourLead': '',
                    'docentsNeeded_Desk': 0,
                    'docentsNeeded_MindfulTour': 0,
                })

        # === CARNEGIE INTERNATIONAL TOURS (cols E=5, F=6) ===
        header_e = ws.cell(start_row, 5).value
        if header_e and 'carnegie international' in str(header_e).lower():
            for r in range(start_row + 1, end_row + 1):
                time_desc = ws.cell(r, 5).value
                docent_name = ws.cell(r, 6).value
                if not time_desc:
                    continue
                time_str = parse_time_from_desc(str(time_desc))
                if not time_str:
                    continue

                name = clean_docent_name(docent_name)
                slot_counter += 1
                sid = f"{date_str}_CI_{slot_counter}"

                is_evening = 'p.m.' in str(time_desc).lower() and any(x in str(time_desc) for x in ['6:', '7:', '6–', '7–'])
                tour_type = 'Carnegie International Evening' if is_evening else 'Carnegie International'

                slots.append({
                    'slotId': sid,
                    'date': date_str,
                    'time': time_str,
                    'tourType': tour_type,
                    'docentsNeeded': 1,
                    'status': 'Assigned' if name else 'Open',
                    'assigned': name or '',
                    'details': '',
                    'tourLeadSchool': '',
                    'participantSchool': '',
                    'mindfulWelcomeDesk': '',
                    'mindfulTourLead': '',
                    'docentsNeeded_Desk': 0,
                    'docentsNeeded_MindfulTour': 0,
                })

        # === SPECIAL TOURS (cols 8+) ===
        for col in range(8, min(ws.max_column + 1, 20)):
            header_val = ws.cell(start_row, col).value
            if not header_val:
                continue
            header_str = str(header_val).strip()
            if not header_str or 'cancelled' in header_str.lower():
                continue

            tour_type, details = classify_tour(header_str)
            if not tour_type:
                continue

            time_str = parse_time_from_tour_desc(header_str)

            # Collect all docent names from rows below in this column
            docent_entries = []
            for r in range(start_row + 1, end_row + 1):
                val = ws.cell(r, col).value
                if val:
                    docent_entries.append(str(val).strip())

            if tour_type == 'Mindful Museum':
                # Parse MM entries into roles
                tour_docents = []
                table_docents = []
                assist_docents = []
                lead_name = ''

                for entry in docent_entries:
                    if is_mm_table(entry):
                        tname = clean_docent_name(entry.replace('MM Table:', '').strip())
                        if tname:
                            table_docents.append(tname)
                    elif is_mm_assist(entry):
                        aname = clean_docent_name(re.sub(r'MM Assist\w*:?\s*', '', entry).strip())
                        if aname:
                            assist_docents.append(aname)
                    else:
                        dname = clean_docent_name(entry)
                        if dname:
                            if is_lead(entry):
                                lead_name = dname
                            tour_docents.append(dname)

                all_names = tour_docents + table_docents + assist_docents
                # Parse actual per-slot docent count from header (e.g. "5 docents")
                mm_count = re.search(r'(\d+)\s*docents', header_str, re.I)
                n_tour = int(mm_count.group(1)) if mm_count else 5
                n_desk = len(table_docents) + len(assist_docents)

                slot_counter += 1
                sid = f"{date_str}_MM_{slot_counter}"
                slots.append({
                    'slotId': sid,
                    'date': date_str,
                    'time': time_str or '10:30 AM',
                    'tourType': 'Mindful Museum',
                    'docentsNeeded': n_tour + n_desk,
                    'status': 'Assigned' if all_names else 'Open',
                    'assigned': ', '.join(all_names) if all_names else '',
                    'details': details or '',
                    'tourLeadSchool': '',
                    'participantSchool': '',
                    'mindfulWelcomeDesk': ', '.join(table_docents + assist_docents) if (table_docents or assist_docents) else '',
                    'mindfulTourLead': ', '.join(tour_docents) if tour_docents else '',
                    'docentsNeeded_Desk': n_desk,
                    'docentsNeeded_MindfulTour': n_tour,
                })

            elif tour_type == 'CI Activation Tour':
                all_names = []
                for entry in docent_entries:
                    name = clean_docent_name(entry)
                    if name:
                        all_names.append(name)

                slot_counter += 1
                sid = f"{date_str}_CIA_{slot_counter}"
                slots.append({
                    'slotId': sid,
                    'date': date_str,
                    'time': time_str or '12:00 PM',
                    'tourType': 'CI Activation Tour',
                    'docentsNeeded': max(3, len(all_names)),
                    'status': 'Assigned' if all_names else 'Open',
                    'assigned': ', '.join(all_names) if all_names else '',
                    'details': details or '',
                    'tourLeadSchool': '',
                    'participantSchool': '',
                    'mindfulWelcomeDesk': '',
                    'mindfulTourLead': '',
                    'docentsNeeded_Desk': 0,
                    'docentsNeeded_MindfulTour': 0,
                })

            elif tour_type in ('School Tour', 'Group Tour'):
                all_names = []
                lead_name = ''
                participants = []
                for entry in docent_entries:
                    name = clean_docent_name(entry)
                    if name:
                        all_names.append(name)
                        if is_lead(entry):
                            lead_name = name
                        else:
                            participants.append(name)

                # Parse group size from details
                size_match = re.search(r'(\d+)\s*(?:to\s*\d+\s*)?(?:adults|students|people|participants)', details or '', re.I)
                group_size = int(size_match.group(1)) if size_match else 0

                needed = max(len(all_names), 1)
                if tour_type == 'School Tour':
                    needed = max(needed, 3)  # school tours typically need more

                slot_counter += 1
                sid = f"{date_str}_GT_{slot_counter}"
                slots.append({
                    'slotId': sid,
                    'date': date_str,
                    'time': time_str or '10:00 AM',
                    'tourType': tour_type,
                    'docentsNeeded': needed,
                    'status': 'Assigned' if all_names else 'Open',
                    'assigned': ', '.join(all_names) if all_names else '',
                    'details': details or '',
                    'tourLeadSchool': lead_name,
                    'participantSchool': ', '.join(participants) if participants else '',
                    'mindfulWelcomeDesk': '',
                    'mindfulTourLead': '',
                    'docentsNeeded_Desk': 0,
                    'docentsNeeded_MindfulTour': 0,
                })


# Sort by date then time
slots.sort(key=lambda s: (s['date'], s['time'] or ''))

# Write CSV
outfile = 'schedule_import.csv'
fields = ['SlotId', 'Date', 'Time', 'TourType', 'DocentsNeeded', 'Status', 'Assigned',
          'Details', 'TourLeadSchool', 'ParticipantSchool', 'MindfulWelcomeDesk',
          'MindfulTourLead', 'DocentsNeeded_Desk', 'DocentsNeeded_MindfulTour']

with open(outfile, 'w', newline='') as f:
    writer = csv.writer(f)
    writer.writerow(fields)
    for s in slots:
        writer.writerow([
            s['slotId'], s['date'], s['time'], s['tourType'], s['docentsNeeded'],
            s['status'], s['assigned'], s['details'], s['tourLeadSchool'],
            s['participantSchool'], s['mindfulWelcomeDesk'], s['mindfulTourLead'],
            s['docentsNeeded_Desk'], s['docentsNeeded_MindfulTour']
        ])

# Summary
n_assigned = sum(1 for s in slots if s['status'] == 'Assigned')
n_open = sum(1 for s in slots if s['status'] == 'Open')
types = {}
for s in slots:
    types[s['tourType']] = types.get(s['tourType'], 0) + 1

print(f"Generated {outfile}: {len(slots)} total slots")
print(f"  Assigned: {n_assigned}")
print(f"  Open: {n_open}")
print(f"  Date range: {slots[0]['date']} to {slots[-1]['date']}")
print(f"\n  Tour types:")
for t, c in sorted(types.items()):
    print(f"    {t}: {c}")

# Show first few and last few
print(f"\n  First 5 slots:")
for s in slots[:5]:
    print(f"    {s['date']} {s['time']} | {s['tourType']} | {s['status']} | {s['assigned']}")
print(f"\n  Last 5 slots:")
for s in slots[-5:]:
    print(f"    {s['date']} {s['time']} | {s['tourType']} | {s['status']} | {s['assigned']}")

# Show unique docent names for verification
all_docents = set()
for s in slots:
    if s['assigned']:
        for n in s['assigned'].split(','):
            n = n.strip()
            if n:
                all_docents.add(n)
print(f"\n  Unique docents found: {len(all_docents)}")
for name in sorted(all_docents):
    print(f"    {name}")
